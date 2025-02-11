# app.py
from flask import Flask, render_template, jsonify, request
from flask_socketio import SocketIO, emit
import os
import openpyxl
from datetime import datetime, timedelta
import threading
import time
import subprocess
import sys

app = Flask(__name__)
socketio = SocketIO(app)

class TimetableApp:
    def __init__(self):
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        self.timetable_dir = os.path.join(self.app_dir, "Jízdní řády")
        self.update_app_dir = os.path.join(self.app_dir, "update_app")
        self.update_script = os.path.join(self.update_app_dir, "updater.py")
        self.timetable = {}
        self.active_sessions = {}  # Sledování aktivních sessions a jejich zastávek

        if not os.path.exists(self.timetable_dir):
            os.makedirs(self.timetable_dir)

    def load_timetables(self):
        xlsx_files = [
            f[:-5] for f in os.listdir(self.timetable_dir)
            if f.endswith(".xlsx") and not f.startswith("~$")
        ]
        xlsx_files.sort(key=lambda name: int(name.split()[0]) if name.split()[0].isdigit() else float('inf'))
        return xlsx_files

    def load_selected_timetable(self, selected_timetable):
        xlsx_path = os.path.join(self.timetable_dir, f"{selected_timetable}.xlsx")
        if os.path.exists(xlsx_path):
            return self.parse_xlsx_timetable(xlsx_path)
        return {}

    def parse_xlsx_timetable(self, xlsx_path):
        timetable = {}
        workbook = openpyxl.load_workbook(xlsx_path)
        for sheet in workbook.sheetnames:
            sheet_data = workbook[sheet]
            stop_name = sheet.strip()
            timetable[stop_name] = []
            for row in sheet_data.iter_rows(min_row=2, max_col=2, values_only=True):
                hour, minutes = row
                if hour is not None and minutes:
                    try:
                        minute_list = [minute.strip() for minute in str(minutes).split(",")]
                        for minute in minute_list:
                            timetable[stop_name].append(f"{int(hour):02}:{minute}")
                    except Exception:
                        continue
        return timetable

    def get_next_times(self, stop_name, timetable_data):
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        times = timetable_data.get(stop_name, [])
        return [time for time in times if time >= current_time]

    def calculate_countdown(self, next_time_str):
        now = datetime.now()
        next_time = datetime.strptime(next_time_str, "%H:%M").replace(
            year=now.year, month=now.month, day=now.day
        )
        if next_time < now:
            next_time += timedelta(days=1)
        diff = next_time - now
        return int(diff.total_seconds() // 60)

timetable_app = TimetableApp()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_lines')
def get_lines():
    lines = timetable_app.load_timetables()
    return jsonify(lines)

@app.route('/get_stops/<line>')
def get_stops(line):
    timetable = timetable_app.load_selected_timetable(line)
    return jsonify(list(timetable.keys()))

@socketio.on('connect')
def handle_connect():
    print('Client connected')

@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')
    if request.sid in timetable_app.active_sessions:
        del timetable_app.active_sessions[request.sid]

@socketio.on('select_stop')
def handle_select_stop(data):
    line = data['line']
    stop = data['stop']
    timetable_data = timetable_app.load_selected_timetable(line)
    next_times = timetable_app.get_next_times(stop, timetable_data)
    
    timetable_app.active_sessions[request.sid] = {
        'line': line,
        'stop': stop
    }
    
    socketio.emit('update_times', {
        'times': next_times,
        'countdown': timetable_app.calculate_countdown(next_times[0]) if next_times else None
    }, room=request.sid)

def background_update():
    while True:
        for sid, session_data in list(timetable_app.active_sessions.items()):
            try:
                line = session_data['line']
                stop = session_data['stop']
                timetable_data = timetable_app.load_selected_timetable(line)
                next_times = timetable_app.get_next_times(stop, timetable_data)
                
                socketio.emit('update_times', {
                    'times': next_times,
                    'countdown': timetable_app.calculate_countdown(next_times[0]) if next_times else None
                }, room=sid)
            except Exception as e:
                print(f"Error updating times for session {sid}: {e}")
        
        time.sleep(60)

if __name__ == '__main__':
    update_thread = threading.Thread(target=background_update, daemon=True)
    update_thread.start()
    socketio.run(app, host='0.0.0.0', debug=True, port=8123)