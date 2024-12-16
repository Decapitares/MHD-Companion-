import os
import openpyxl
from pywebio import start_server
from pywebio.output import put_buttons, put_text, put_scrollable, clear, toast, use_scope, put_html
from pywebio.session import register_thread, get_current_session, SessionNotFoundException, eval_js
from pywebio.input import select
from datetime import datetime
import time
import threading
from datetime import datetime, timedelta
import subprocess
import sys

def get_corrected_time():
    """Získá aktuální čas od klienta a přidá +1 hodinu pro korekci."""
    js_code = """
    let now = new Date();
    now.toISOString();
    """
    client_time_str = eval_js(js_code)
    client_time = datetime.fromisoformat(client_time_str[:-1])
    corrected_time = client_time + timedelta(hours=1)

    return corrected_time

class TimetableApp:
    def __init__(self):
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        self.timetable_dir = os.path.join(self.app_dir, "Jízdní řády")

        # Cesta k updater skriptu
        self.update_app_dir = os.path.join(self.app_dir, "update_app")
        self.update_script = os.path.join(self.update_app_dir, "updater.py")

        self.timetable = {}
        self.current_stop = None
        self.current_times = []
        self.current_section = "stops"
        self.running = True

        if not os.path.exists(self.timetable_dir):
            os.makedirs(self.timetable_dir)

    def start_countdown_thread(self):
        """Spustí vlákno pro dynamický odpočet a registruje ho."""
        countdown_thread = threading.Thread(target=self.update_countdown, daemon=True)
        register_thread(countdown_thread)
        countdown_thread.start()

    def load_timetables(self):
        xlsx_files = [
            f[:-5] for f in os.listdir(self.timetable_dir)
            if f.endswith(".xlsx") and not f.startswith("~$")
        ]
        xlsx_files.sort(key=lambda name: int(name.split()[0]) if name.split()[0].isdigit() else float("inf"))
        return ["Vyberte si linku"] + xlsx_files

    def load_selected_timetable(self, selected_timetable):
        xlsx_path = os.path.join(self.timetable_dir, f"{selected_timetable}.xlsx")
        if os.path.exists(xlsx_path):
            self.timetable = self.parse_xlsx_timetable(xlsx_path)
        else:
            self.timetable = {}

    def parse_xlsx_timetable(self, xlsx_path):
        timetable = {}
        workbook = openpyxl.load_workbook(xlsx_path)
        for sheet in workbook.sheetnames:
            sheet_data = workbook[sheet]
            stop_name = sheet.strip()
            timetable[stop_name] = []
            for row in sheet_data.iter_rows(min_row=2, max_col=2, values_only=True):
                hour, minutes = row
                if hour is not None and minutes:
                    try:
                        minute_list = [minute.strip() for minute in str(minutes).split(",")]
                        for minute in minute_list:
                            timetable[stop_name].append(f"{int(hour):02}:{minute}")
                    except Exception:
                        pass
        return timetable

    def render_ui(self):
        clear()
        if self.current_section == "stops":
            self.render_stops_section()
        elif self.current_section == "times":
            self.render_times_section()
        self.render_navigation_buttons()
        self.render_update_button()  # Přidání tlačítka Update data

    def render_stops_section(self):
        put_text("Vyberte linku:")
        timetable_list = self.load_timetables()
        selected_timetable = select("Linka:", timetable_list)
        if selected_timetable and selected_timetable != "Vyberte si linku":
            self.load_selected_timetable(selected_timetable)

        put_text("Vyberte zastávku:")
        stop_buttons = [{"label": stop, "value": stop} for stop in self.timetable.keys()]
        put_buttons(stop_buttons, onclick=self.show_times)

    def render_times_section(self):
        if self.current_stop:
            with use_scope("countdown_scope", clear=True):
                put_text("Příjezd za: ...").style('font-size: 20px; font-weight: bold;')
            put_text(f"Časy příjezdů pro zastávku {self.current_stop}:")
            self.update_times()
            self.start_countdown_thread()
        else:
            put_text("Nejprve vyberte zastávku v sekci 'Zastávky'.")

    def update_times(self):
        try:
            corrected_time = get_corrected_time()
            self.current_times = [
                time for time in self.timetable.get(self.current_stop, [])
                if time >= corrected_time.strftime("%H:%M")
            ]
            with use_scope("times_scope", clear=True):
                if self.current_times:
                    put_scrollable("\n".join(self.current_times), height=200)
                else:
                    put_text("Žádné další příjezdy dnes.")
        except Exception as e:
            print(f"Chyba při aktualizaci časů: {e}")

    def update_countdown(self):
        try:
            while self.running:
                corrected_time = get_corrected_time()

                if self.current_stop and self.current_times:
                    self.current_times = [
                        time for time in self.timetable.get(self.current_stop, [])
                        if time >= corrected_time.strftime("%H:%M")
                    ]

                    if not self.current_times:
                        with use_scope("countdown_scope", clear=True):
                            put_text("Žádné další příjezdy dnes.").style('font-size: 20px; font-weight: bold;')
                        break

                    next_time_str = self.current_times[0]
                    next_time_naive = datetime.strptime(next_time_str, "%H:%M")
                    next_time = corrected_time.replace(hour=next_time_naive.hour, minute=next_time_naive.minute, second=0, microsecond=0)
                    minutes_to_next = int((next_time - corrected_time).total_seconds() // 60)

                    with use_scope("countdown_scope", clear=True):
                        put_text(f"Příjezd za: {minutes_to_next} min").style('font-size: 20px; font-weight: bold;')

                    with use_scope("times_scope", clear=True):
                        put_scrollable("\n".join(self.current_times), height=200)
                else:
                    with use_scope("countdown_scope", clear=True):
                        put_text("Žádné další příjezdy dnes.").style('font-size: 20px; font-weight: bold;')
                    break

                time.sleep(60)
        except SessionNotFoundException:
            self.running = False

    def render_navigation_buttons(self):
        put_buttons([
            {"label": "Zpět", "value": "stops", "style": "background-color: green; color: white; font-size: 18px; margin-right: 20px;"},
        ], onclick=self.navigate)

    def render_update_button(self):
        put_html("<hr>")
        put_buttons([{"label": "Update data", "value": "update"}], onclick=self.run_update_script)

    def run_update_script(self, _):
        try:
            subprocess.run([sys.executable, self.update_script], check=True)
            toast("Data byla úspěšně aktualizována.")
            # Po aktualizaci můžeme třeba znovu načíst rozhraní
            self.render_ui()
        except Exception as e:
            toast(f"Nepodařilo se aktualizovat data: {e}")

    def navigate(self, section):
        self.current_section = section
        self.render_ui()

    def show_times(self, stop):
        self.current_stop = stop
        self.current_section = "times"
        self.start_countdown_thread()
        self.render_ui()

    def start(self):
        self.render_ui()

def main():
    app = TimetableApp()
    app.start()

if __name__ == "__main__":
    start_server(main, port=5000, debug=True)
