import os
import openpyxl
from pywebio import start_server
from pywebio.output import put_buttons, put_text, put_scrollable, clear, toast, use_scope
from pywebio.session import register_thread, get_current_session, SessionNotFoundException, eval_js  # Přidáme import
from pywebio.input import select
from datetime import datetime
import time
import threading
from pytz import timezone, utc

import ntplib
from datetime import datetime, timezone, timedelta

def get_corrected_time():
    """Získá aktuální čas a přidá +1 hodinu pro korekci."""
    js_code = """
    let now = new Date();
    now.toISOString();  // Vrátí čas jako ISO string
    """
    client_time_str = eval_js(js_code)  # Získáme ISO string času z klienta
    client_time = datetime.fromisoformat(client_time_str[:-1])  # Převedeme na datetime objekt

    # Přičteme 1 hodinu
    corrected_time = client_time + timedelta(hours=1)

    # Debug: Výpis původního a opraveného času
    print(f"Původní čas klienta: {client_time}")
    print(f"Opravený čas klienta: {corrected_time}")

    return corrected_time

class TimetableApp:
    def __init__(self):
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        self.timetable_dir = os.path.join(self.app_dir, "Jízdní řády")
        self.timetable = {}
        self.current_stop = None
        self.current_times = []
        self.current_section = "stops"  # Výchozí sekce
        self.running = True  # Kontrola běhu odpočtu

        if not os.path.exists(self.timetable_dir):
            os.makedirs(self.timetable_dir)
            
    def start_countdown_thread(self):
        """Spustí vlákno pro dynamický odpočet a registruje ho."""
        countdown_thread = threading.Thread(target=self.update_countdown, daemon=True)
        register_thread(countdown_thread)  # Zaregistrujeme vlákno
        countdown_thread.start()  # Spustíme vlákno

    def load_timetables(self):
        xlsx_files = [
            f[:-5] for f in os.listdir(self.timetable_dir)
            if f.endswith(".xlsx") and not f.startswith("~$")
        ]
        xlsx_files.sort(key=lambda name: int(name.split()[0]) if name.split()[0].isdigit() else float("inf"))
        return ["Vyberte si linku"] + xlsx_files

    def load_selected_timetable(self, selected_timetable):
        xlsx_path = os.path.join(self.timetable_dir, f"{selected_timetable}.xlsx")
        if os.path.exists(xlsx_path):
            self.timetable = self.parse_xlsx_timetable(xlsx_path)
        else:
            self.timetable = {}

    def parse_xlsx_timetable(self, xlsx_path):
        timetable = {}
        workbook = openpyxl.load_workbook(xlsx_path)
        for sheet in workbook.sheetnames:
            sheet_data = workbook[sheet]
            stop_name = sheet.strip()
            timetable[stop_name] = []
            for row in sheet_data.iter_rows(min_row=2, max_col=2, values_only=True):
                hour, minutes = row
                if hour is not None and minutes:
                    try:
                        minute_list = [minute.strip() for minute in str(minutes).split(",")]
                        for minute in minute_list:
                            timetable[stop_name].append(f"{int(hour):02}:{minute}")
                    except Exception as e:
                        pass
        return timetable

    def render_ui(self):
        clear()
        if self.current_section == "stops":
            self.render_stops_section()
        elif self.current_section == "times":
            self.render_times_section()
        self.render_navigation_buttons()

    def render_stops_section(self):
        put_text("Vyberte linku:")
        timetable_list = self.load_timetables()
        selected_timetable = select("Linka:", timetable_list)
        if selected_timetable and selected_timetable != "Vyberte si linku":
            self.load_selected_timetable(selected_timetable)

        put_text("Vyberte zastávku:")
        stop_buttons = [{"label": stop, "value": stop} for stop in self.timetable.keys()]
        put_buttons(stop_buttons, onclick=self.show_times)

    def render_times_section(self):
        """Vykreslí sekci časů příjezdů."""
        if self.current_stop:
            with use_scope("countdown_scope", clear=True):
                # Umístíme placeholder pro odpočet (bude aktualizován vlákny)
                put_text("Příjezd za: ...").style('font-size: 20px; font-weight: bold;')
            put_text(f"Časy příjezdů pro zastávku {self.current_stop}:")
            self.update_times()  # Zobrazení všech časů
            self.start_countdown_thread()  # Spuštění vlákna pro dynamický odpočet
        else:
            put_text("Nejprve vyberte zastávku v sekci 'Zastávky'.")
    
    def update_times(self):
        """Aktualizuje a zobrazuje časy příjezdů podle aktuálního času."""
        try:
            # Získáme opravený čas
            corrected_time = get_corrected_time()
    
            # Debug: Výpis aktuálního opraveného času
            print(f"Aktuální opravený čas: {corrected_time}")
    
            # Filtrování časů, aby zůstaly jen budoucí časy
            self.current_times = [
                time for time in self.timetable.get(self.current_stop, [])
                if time >= corrected_time.strftime("%H:%M")
            ]
    
            # Debug: Výpis aktuálních časů po filtrování
            print(f"Časy příjezdů po aktualizaci: {self.current_times}")
    
            # Aktualizace zobrazení časů
            with use_scope("times_scope", clear=True):
                if self.current_times:
                    put_scrollable(
                        "\n".join(self.current_times), height=200
                    )
                else:
                    put_text("Žádné další příjezdy dnes.")
        except Exception as e:
            print(f"Chyba při aktualizaci časů: {e}")
    
    def update_countdown(self):
        """Pravidelně aktualizuje odpočet podle opraveného času."""
        try:
            while self.running:
                # Získáme opravený čas
                corrected_time = get_corrected_time()
    
                # Debug: Výpis aktuálního opraveného času
                print(f"Aktuální opravený čas v odpočtu: {corrected_time}")
    
                if self.current_stop and self.current_times:
                    # Dynamicky aktualizujeme seznam aktuálních časů
                    self.current_times = [
                        time for time in self.timetable.get(self.current_stop, [])
                        if time >= corrected_time.strftime("%H:%M")
                    ]
    
                    # Debug: Výpis aktuálních časů po aktualizaci
                    print(f"Časy příjezdů po aktualizaci: {self.current_times}")
    
                    # Pokud nejsou další časy, zobrazíme zprávu a ukončíme odpočet
                    if not self.current_times:
                        with use_scope("countdown_scope", clear=True):
                            put_text("Žádné další příjezdy dnes.").style('font-size: 20px; font-weight: bold;')
                        break
    
                    # Výpočet odpočtu k nejbližšímu času
                    next_time_str = self.current_times[0]
                    next_time_naive = datetime.strptime(next_time_str, "%H:%M")
                    next_time = corrected_time.replace(hour=next_time_naive.hour, minute=next_time_naive.minute, second=0, microsecond=0)
    
                    # Výpočet minut do dalšího příjezdu
                    minutes_to_next = int((next_time - corrected_time).total_seconds() // 60)
    
                    # Aktualizace odpočtu a seznamu časů
                    with use_scope("countdown_scope", clear=True):
                        put_text(f"Příjezd za: {minutes_to_next} min").style('font-size: 20px; font-weight: bold;')
    
                    # Aktualizace seznamu časů
                    with use_scope("times_scope", clear=True):
                        put_scrollable(
                            "\n".join(self.current_times), height=200
                        )
                else:
                    # Pokud nejsou vybrány časy nebo zastávka
                    with use_scope("countdown_scope", clear=True):
                        put_text("Žádné další příjezdy dnes.").style('font-size: 20px; font-weight: bold;')
                    break
    
                time.sleep(60)  # Aktualizace každou minutu
        except SessionNotFoundException:
            self.running = False

    def render_navigation_buttons(self):
        put_buttons([
            {"label": "Zpět1", "value": "stops", "style": "background-color: green; color: white; font-size: 18px; margin-right: 20px;"},
        ], onclick=self.navigate)


    def navigate(self, section):
        self.current_section = section
        self.render_ui()

    def show_times(self, stop):
        """Uloží vybranou zastávku a přepne na sekci časů."""
        self.current_stop = stop
        self.current_section = "times"
        self.start_countdown_thread()  # Spuštění registrovaného vlákna
        self.render_ui()

    def start(self):
        self.render_ui()

def main():
    app = TimetableApp()
    app.start()

if __name__ == "__main__":
    start_server(main, port=5000, debug=True)
